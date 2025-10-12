#!/usr/bin/env python3
"""
Convert ASHG2025 schedule markdown to Excel spreadsheet
"""
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def parse_schedule_markdown(md_path):
    """Parse schedule markdown and extract structured data"""

    with open(md_path, 'r', encoding='utf-8') as f:
        content = f.read()

    talks = []
    current_day = None

    # Split into sections by day
    day_pattern = r'^## 📅 (.+)$'
    section_pattern = r'^### (🔴 )?(.+?)$\n\n\*\*Type:\*\* (.+?) \| \*\*Relevance Score:\*\* (.+?)$\n\n\*\*⏰ Time:\*\* (.+?)$\n\n\*\*👥 Authors:\*\* (.+?)$\n\n\*\*📝 Abstract:\*\*\n\n(.+?)(?:\n\n⚠️ \*\*CONFLICT:\*\* (.+?))?\n\n---'

    lines = content.split('\n')
    i = 0

    while i < len(lines):
        line = lines[i]

        # Check for day header
        day_match = re.match(day_pattern, line)
        if day_match:
            current_day = day_match.group(1)
            i += 1
            continue

        # Check for talk section
        if line.startswith('### '):
            # Extract conflict marker
            has_conflict = '🔴' in line
            title = line.replace('### 🔴 ', '').replace('### ', '')

            # Get next lines
            i += 1
            if i >= len(lines):
                break
            i += 1  # Skip blank line

            # Type and relevance
            if i < len(lines) and '**Type:**' in lines[i]:
                type_line = lines[i]
                type_match = re.search(r'\*\*Type:\*\* (.+?) \| \*\*Relevance Score:\*\* (.+?)$', type_line)
                if type_match:
                    talk_type = type_match.group(1).replace('🎤 ', '').replace('📋 ', '')
                    relevance = type_match.group(2)
                else:
                    talk_type = "Unknown"
                    relevance = "N/A"
                i += 1
            else:
                talk_type = "Unknown"
                relevance = "N/A"

            i += 1  # Skip blank line

            # Time
            if i < len(lines) and '**⏰ Time:**' in lines[i]:
                time_match = re.search(r'\*\*⏰ Time:\*\* (.+?)$', lines[i])
                time = time_match.group(1) if time_match else "TBD"
                i += 1
            else:
                time = "TBD"

            i += 1  # Skip blank line

            # Authors
            if i < len(lines) and '**👥 Authors:**' in lines[i]:
                authors_match = re.search(r'\*\*👥 Authors:\*\* (.+?)$', lines[i])
                authors = authors_match.group(1) if authors_match else "Unknown"
                i += 1
            else:
                authors = "Unknown"

            i += 1  # Skip blank line

            # Abstract header
            if i < len(lines) and '**📝 Abstract:**' in lines[i]:
                i += 1
                i += 1  # Skip blank line

            # Abstract content
            abstract_lines = []
            while i < len(lines) and lines[i].strip() and not lines[i].startswith('⚠️') and not lines[i].startswith('---'):
                abstract_lines.append(lines[i])
                i += 1
            abstract = ' '.join(abstract_lines)

            # Check for conflict note
            conflict_note = ""
            if i < len(lines) and lines[i].startswith('⚠️'):
                conflict_match = re.search(r'⚠️ \*\*CONFLICT:\*\* (.+?)$', lines[i])
                if conflict_match:
                    conflict_note = conflict_match.group(1)
                i += 1

            # Add to talks list
            talks.append({
                'Day': current_day,
                'Time': time,
                'Title': title,
                'Type': talk_type,
                'Relevance': relevance,
                'Authors': authors,
                'Abstract': abstract,
                'Has Conflict': 'Yes' if has_conflict else 'No',
                'Conflict Note': conflict_note
            })

        i += 1

    return talks

def create_excel(talks, output_path):
    """Create formatted Excel file from talks data"""

    # Create DataFrame
    df = pd.DataFrame(talks)

    # Convert relevance to numeric for sorting (handle N/A values)
    df['Relevance_Numeric'] = df['Relevance'].str.rstrip('%').replace('N/A', '0').astype(float)
    df = df.sort_values(['Day', 'Time', 'Relevance_Numeric'], ascending=[True, True, False])
    df = df.drop('Relevance_Numeric', axis=1)

    # Write to Excel
    df.to_excel(output_path, index=False, sheet_name='ASHG2025 Schedule')

    # Load workbook for formatting
    wb = load_workbook(output_path)
    ws = wb.active

    # Header formatting
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Column widths
    column_widths = {
        'A': 20,  # Day
        'B': 20,  # Time
        'C': 50,  # Title
        'D': 10,  # Type
        'E': 12,  # Relevance
        'F': 30,  # Authors
        'G': 70,  # Abstract
        'H': 12,  # Has Conflict
        'I': 30   # Conflict Note
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Row formatting
    conflict_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')

    for row in range(2, ws.max_row + 1):
        # Wrap text and align
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Highlight conflicts
        if ws.cell(row=row, column=8).value == 'Yes':  # Has Conflict column
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = conflict_fill

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save formatted workbook
    wb.save(output_path)
    print(f"✅ Excel file created: {output_path}")

def main():
    md_path = '/Users/camellia/PycharmProjects/PhD_Agent/conference/ASHG2025/ashg2025_schedule.md'
    output_path = '/Users/camellia/PycharmProjects/PhD_Agent/conference/ASHG2025/ashg2025_schedule.xlsx'

    print("📄 Parsing markdown schedule...")
    talks = parse_schedule_markdown(md_path)
    print(f"   Found {len(talks)} talks")

    print("📊 Creating Excel spreadsheet...")
    create_excel(talks, output_path)

    print(f"\n🎉 Conversion complete!")
    print(f"   Output: {output_path}")

if __name__ == '__main__':
    main()
